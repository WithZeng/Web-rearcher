"""Write extracted data to CSV, Excel, JSON, and BibTeX."""

from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path

from .config import FIELDS, OUTPUT_DIR

_FIELD_CN = {
    "gelma_concentration": "GelMA浓度/concentration(%)",
    "degree_of_substitution": "取代度(接枝率)/degree of substitution",
    "gelma_molecular_weight": "GelMA分子量/KDa",
    "microsphere_size": "微球粒径分布/μm",
    "drug_microsphere_ratio": "药物微球质量比(药物/微球)",
    "encapsulation_efficiency": "包封率/%",
    "drug_loading_rate": "载药率/%",
    "drug_loading_amount": "载药量",
    "drug_name": "药物名称",
    "drug_molecular_weight": "药物分子量",
    "tpsa": "拓扑极性表面积TPSA",
    "hbd": "氢键供体数(HBD)",
    "hba": "氢键受体数(HBA)",
    "drug_nha": "杂原子数量Drug_NHA",
    "drug_melting_point": "熔点Drug_Tm",
    "pka": "酸解离常数(pKa)",
    "drug_logp": "计算分配系数Drug_LogP",
    "temperature": "温度/°C",
    "ph": "pH",
    "release_time": "释放时间/h",
    "release_amount": "释放量/%release",
    "source_title": "文献来源",
    "source_doi": "DOI",
}

_META_COLUMNS = ["_data_quality", "text_source", "_review", "_review_score", "_review_flags", "_quality_label", "_quality_total"]
_META_CN = {
    "_data_quality": "数据质量",
    "text_source": "文本来源",
    "_review": "审查结果",
    "_review_score": "审查分数",
    "_review_flags": "审查问题",
    "_quality_label": "质量等级",
    "_quality_total": "质量总分",
}

ALL_COLUMNS = FIELDS + _META_COLUMNS
ALL_CN = {**_FIELD_CN, **_META_CN}

_GROUP_HEADER = [
    ("GelMA微球", 8),
    ("药物特征", 9),
    ("环境特征", 3),
    ("目标量", 1),
    ("文献来源", 2),
    ("元数据", 7),
]


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def write_csv(rows: list[dict], path: Path | None = None) -> Path:
    path = path or OUTPUT_DIR / f"results_{_timestamp()}.csv"
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=ALL_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    return path


def write_excel(rows: list[dict], path: Path | None = None) -> Path:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path = path or OUTPUT_DIR / f"results_{_timestamp()}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    quality_fill_good = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    quality_fill_mid = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    quality_fill_low = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    col = 1
    for group_name, span in _GROUP_HEADER:
        ws.cell(row=1, column=col, value=group_name).font = header_font
        ws.cell(row=1, column=col).alignment = center
        if span > 1:
            ws.merge_cells(
                start_row=1, start_column=col,
                end_row=1, end_column=col + span - 1,
            )
        col += span

    for i, field in enumerate(ALL_COLUMNS, start=1):
        cell = ws.cell(row=2, column=i, value=ALL_CN.get(field, field))
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    quality_col_idx = ALL_COLUMNS.index("_data_quality") + 1

    for row_idx, row in enumerate(rows, start=3):
        for col_idx, f in enumerate(ALL_COLUMNS, start=1):
            val = row.get(f)
            ws.cell(row=row_idx, column=col_idx, value=val)

        quality_val = row.get("_data_quality")
        if quality_val is not None:
            cell = ws.cell(row=row_idx, column=quality_col_idx)
            cell.number_format = "0%"
            if quality_val >= 0.7:
                cell.fill = quality_fill_good
            elif quality_val >= 0.4:
                cell.fill = quality_fill_mid
            else:
                cell.fill = quality_fill_low

    for i, field in enumerate(ALL_COLUMNS, start=1):
        label = ALL_CN.get(field, field)
        ws.column_dimensions[get_column_letter(i)].width = max(len(label) + 4, 12)

    wb.save(path)
    return path


def write_json(rows: list[dict], path: Path | None = None) -> Path:
    """Write rows as a JSON array."""
    path = path or OUTPUT_DIR / f"results_{_timestamp()}.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    return path


def write_bibtex(rows: list[dict], path: Path | None = None) -> Path:
    """Write rows as BibTeX entries for import into Zotero/EndNote."""
    path = path or OUTPUT_DIR / f"results_{_timestamp()}.bib"
    lines: list[str] = []
    for i, row in enumerate(rows):
        doi = row.get("source_doi", "")
        title = row.get("source_title", "")
        drug = row.get("drug_name", "")
        key = doi.replace("/", "_").replace(".", "_") if doi else f"paper_{i+1}"

        entry = [f"@article{{{key},"]
        if title:
            entry.append(f"  title = {{{title}}},")
        if doi:
            entry.append(f"  doi = {{{doi}}},")
        if drug:
            entry.append(f"  keywords = {{{drug}}},")
        entry.append("}")
        lines.append("\n".join(entry))

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n\n".join(lines) + "\n")
    return path


def filter_rows_by_quality(
    rows: list[dict],
    min_review: str = "ok",
) -> list[dict]:
    """Filter rows by review status. Keeps 'ok' and optionally 'low_quality'."""
    allowed = {"ok"}
    if min_review == "low_quality":
        allowed.add("low_quality")
    elif min_review == "all":
        return rows
    return [r for r in rows if r.get("_review", "ok") in allowed]


_CORE_FIELDS = [
    "drug_name", "gelma_concentration", "microsphere_size",
    "encapsulation_efficiency", "release_amount", "release_time",
]


def _safe_float(value: object, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def filter_empty_rows(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """Remove rows with no useful extracted data.
    
    Returns (kept, discarded).
    """
    kept, discarded = [], []
    for r in rows:
        quality = r.get("_data_quality")
        q = _safe_float(quality)
        if q <= 0:
            has_core = any(r.get(f) for f in _CORE_FIELDS)
            if not has_core:
                discarded.append(r)
                continue
        text_src = r.get("text_source", "none")
        if text_src == "none":
            has_core = any(r.get(f) for f in _CORE_FIELDS)
            if not has_core:
                discarded.append(r)
                continue
        kept.append(r)
    return kept, discarded
